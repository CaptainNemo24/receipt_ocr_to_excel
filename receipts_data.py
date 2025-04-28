import os
import requests
import uuid
import time
import json
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# CLOVA OCR API 호출
api_url = 'YOUR_API_URL'
secret_key = 'YOUR_SECRET_KEY'
# 이스케이프 문자 무시 + 포맷 가능성을 위해 fr를 통해 파일 경로 읽기
image_file = fr'YOUR_FILE_FATH\jpg'

request_json = {
    'images': [
        {
            'format': 'jpg',
            'name': 'demo'
        }
    ],
    'requestId': str(uuid.uuid4()),
    'version': 'V2',
    'timestamp': int(round(time.time() * 1000))
}

payload = {'message': json.dumps(request_json).encode('UTF-8')}

# 대량 image file 처리와 원하는 경로에서 파일 로드 및 생성
for filename in os.listdir(image_file):
    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
        image_path = os.path.join(image_file, filename)
        print(f"[{filename}] 처리 중...")

        with open(image_path, 'rb') as f:
            files = [('file', f)]
            headers = {
            'X-OCR-SECRET': secret_key
            }
            response = requests.request("POST", api_url, headers=headers, data = payload, files = files)
            json_data = fr"YOUR_FILE_FATH\json\{filename}".replace(".jpg", ".json")

        # json 파일이 있을 경우에는 불러오고, 없을 경우에는 새로 생성
        try:
            with open(json_data, 'r', encoding='utf-8') as f:
                data = json.load(f)
                fields = data['images'][0]['fields']
                print(f"{json_data} 로드 완료")
        except FileNotFoundError:
            with open(json_data, 'w', encoding='utf-8') as f:
                json.dump(response.json(), f, ensure_ascii=False, indent=2)
                fields = response.json()['images'][0]['fields']
                print(f"{json_data} 생성 완료")
        
        # 생성된 json 파일에서 text만 추출(이모티콘, 그림 등 제외)
        extract_text = [f['inferText'] for f in fields]

        # 영수증 패턴을 분석하여 품목 추출을 위한 정규 표현식
        def is_product_name(s):
            return re.match(r'\*[\w가-힣]+', s)
        
        items = []
        store_name = None
        sales_date = None

        # 매장명 / 매출일 추출
        for w, text in enumerate(extract_text):
            if text == "[매장명]" and w + 1 < len(extract_text):
                store_name = extract_text[w + 1]
            elif text == "[매출일]" and w + 1 < len(extract_text):
                sale_date = extract_text[w + 1]
                sales_date = datetime.strptime(sale_date, "%Y-%m-%d").date()

        # 상품 정보 추출
        i = 0
        while i + 3 < len(extract_text):
            if is_product_name(extract_text[i]):
                product_name = extract_text[i]
                unit_price = extract_text[i + 1]
                quantity = extract_text[i + 2]
                amount = extract_text[i + 3]

                items.append({
                        "날짜": sales_date,
                        "업체명": store_name,
                        "품목": product_name,
                        "단가": unit_price,
                        "수량": quantity,
                        "금액": amount
                })
                i += 4
            else:
                i += 1

        # "*부가세" 품목만 제거
        items = [item for item in items if item["품목"] != "*부가세"]

        # 결과 출력
        for item in items:
            item["품목"] = item["품목"].replace("*", "")
            for key in ["단가", "수량", "금액"]:
                # 쉼표나 공백 등 제거하고 숫자로 변환
                cleaned = item[key].replace(",", "")
                item[key] = int(cleaned)
            print(item)

        # 데이터 프레임으로 전환 및 생성
        receipt_data = pd.DataFrame(items)
        
        # 파일 경로 및 오픈할 sheet 이름
        file_path = fr"YOUR_FILE_FATH\csv\샘플 데이터.xlsx"
        sheet_name = "지출내역"

        # 엑셀 열기
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # 테이블 정보 가져오기
        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]
        start_cell, end_cell = table.ref.split(":")
        start_col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col_letter = ''.join(filter(str.isalpha, end_cell))

        start_col_index = column_index_from_string(start_col_letter)
        end_col_index = column_index_from_string(end_col_letter)

        # 실제 데이터가 있는 마지막 행 찾기
        def get_last_data_row(ws, start_row, col_index):
            row = start_row
            while ws.cell(row=row, column=col_index).value:
                row += 1
            return row - 1

        actual_last_row = get_last_data_row(ws, start_row + 1, start_col_index)

        # 새 데이터 삽입 (실제 마지막 데이터 아래부터)
        for r_idx, row in enumerate(receipt_data.values.tolist(), start=actual_last_row + 1):
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx, column=start_col_index + c_idx, value=value)

        # 테이블 범위 재설정
        new_end_row = actual_last_row + len(receipt_data)
        new_ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_end_row}"

        # 지정한 파일 경로에 저장 및 확인
        wb.save(file_path)
        print(f"{file_path} 파일에 저장되었습니다.")
